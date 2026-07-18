"""
employee_import.py — Employee Excel Import module
Save to: database/routes/employee_import.py
"""

import io
import re
import logging
from datetime import datetime, date

from flask import Blueprint, request, jsonify, send_file, session
from flask_login import login_required, current_user
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models import db, Employee, ProfessionMaster, BuyerMaster, EmployeeProfession

emp_import_bp = Blueprint('emp_import', __name__)
log = logging.getLogger('employee_import')

COLUMNS = [
    ('Employee Code', 'employee_code'),
    ('Auto Code', 'auto_code'),
    ('Active', 'is_active'),
    ('Muslim', 'is_muslim'),
    ('Employee Name', 'name'),
    ('Employee Name Arabic', 'name_ar'),
    ('Nationality', 'nationality'),
    ('Nationality Arabic', 'nationality_ar'),
    ('Profession', 'profession'),
    ('Profession Arabic', 'profession_ar'),
    ('Education', 'education'),
    ('Education Arabic', 'education_ar'),
    ('Kafeel Name', 'kafeel_name'),
    ('Kafeel Name Arabic', 'kafeel_name_ar'),
    ('Kafeel Reference', 'kafeel_reference'),
    ('Kafeel Reference Arabic', 'kafeel_reference_ar'),
    ('Kafalat Number', 'kafalat_number'),
    ('Passport Number', 'passport_number'),
    ('Passport Expiry', 'passport_expiry'),
    ('Passport Location', 'passport_location'),
    ('Entry Number', 'entry_number'),
    ('Iqama Number', 'iqama_number'),
    ('Iqama Expiry', 'iqama_expiry'),
    ('Arrival Date', 'arrival_date'),
    ('Birth Date', 'birth_date'),
    ('Joining Date', 'joining_date'),
    ('End Date Work', 'end_date_work'),
    ('Mobile', 'mobile'),
    ('Email', 'email'),
    ('Address', 'address'),
    ('Address Arabic', 'address_ar'),
    ('Home City', 'home_city'),
    ('Home City Arabic', 'home_city_ar'),
    ('Employee Reference', 'employee_reference'),
    ('Employee Reference Arabic', 'employee_reference_ar'),
    ('Company', 'company'),
    ('Company Arabic', 'company_ar'),
    ('Department', 'department'),
    ('Department Arabic', 'department_ar'),
    ('Section', 'section'),
    ('Section Arabic', 'section_ar'),
    ('Work Month', 'work_month'),
    ('Work Status', 'work_status'),
    ('Shift Type', 'shift_type'),
    ('Salary Type', 'salary_type'),
    ('Basic Salary', 'basic_salary'),
    ('Total Allowances', 'total_allowances'),
    ('Net Salary', 'net_salary'),
    ('PO Number', 'po_number'),
    ('PO Rate', 'po_rate'),
    ('Working Hours', 'working_hours'),
    ('Overtime Ratio', 'overtime_ratio'),
    ('Overtime Rate', 'overtime_rate'),
    ('Hostel Name', 'hostel_name'),
    ('Hostel Name Arabic', 'hostel_name_ar'),
    ('Room Number', 'room_number'),
    ('Hostel Location', 'hostel_location'),
    ('Hostel Location Arabic', 'hostel_location_ar'),
    ('CRN', 'crn'),
    ('CRN Arabic', 'crn_ar'),
    ('Insurance Company', 'insurance_company'),
    ('Insurance Company Arabic', 'insurance_company_ar'),
    ('Insurance Expiry', 'insurance_expiry'),
    ('Labour Office', 'labour_office'),
    ('Buyer', 'buyer'),
]

HEADERS = [c[0] for c in COLUMNS]
FIELD_BY_HDR = dict(COLUMNS)

# Only Name and Joining Date are required — employee_code is auto-generated if blank
REQUIRED = {'name', 'joining_date'}
DATE_FLDS = {'passport_expiry', 'iqama_expiry', 'arrival_date', 'birth_date',
             'joining_date', 'end_date_work', 'insurance_expiry'}
NUM_FLDS = {'basic_salary', 'total_allowances', 'net_salary', 'po_rate',
            'working_hours', 'overtime_ratio', 'overtime_rate'}
BOOL_FLDS = {'auto_code', 'is_active', 'is_muslim'}

EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
MOBILE_RE = re.compile(r'^\+?[\d\s\-\(\)]{7,20}$')

ALLOWED_EXT = {'.xlsx', '.xls'}


def _t(en, ar):
    return ar if session.get('lang') == 'ar' else en


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_admin():
            return jsonify({'ok': False, 'error': _t('Access denied.', 'الوصول مرفوض')}), 403
        return f(*args, **kwargs)
    return decorated


def generate_emp_code():
    """Auto-generate employee code: EMP-{last_id+1:04d}"""
    last = Employee.query.order_by(Employee.id.desc()).first()
    num = (last.id + 1) if last else 1
    return f'EMP-{num:04d}'


class RowError(Exception):
    pass


def parse_date_cell(value, field_label):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise RowError(f'Invalid date format in "{field_label}": {s}')


def parse_number_cell(value, field_label):
    if value is None or value == '':
        return 0
    try:
        return float(str(value).strip().replace(',', ''))
    except (ValueError, TypeError):
        raise RowError(f'Invalid numeric value in "{field_label}": {value}')


def parse_bool_cell(value, field_label, default=False):
    if value is None or value == '':
        return default
    s = str(value).strip().lower()
    if s in ('yes', 'true', '1', 'y', 'نعم'):
        return True
    if s in ('no', 'false', '0', 'n', 'لا'):
        return False
    raise RowError(f'Invalid boolean value in "{field_label}": {value}')


def clean_str(value):
    if value is None:
        return ''
    return str(value).strip()


def validate_email(value):
    v = clean_str(value)
    if v and not EMAIL_RE.match(v):
        raise RowError(f'Invalid email: {v}')
    return v


def validate_mobile(value):
    v = clean_str(value)
    if v and not MOBILE_RE.match(v):
        raise RowError(f'Invalid mobile number: {v}')
    return v


def lookup_profession_ids(prof_str, lookups):
    """
    Parse comma-separated or single profession string.
    Returns (profession_ids_list, first_profession_name, first_profession_ar_name, error_message_or_None)
    """
    if not prof_str:
        return [], '', '', None

    names = [n.strip() for n in prof_str.split(',') if n.strip()]
    if not names:
        return [], '', '', None

    ids = []
    first_en = ''
    first_ar = ''
    for name in names:
        pid = lookups['professions'].get(name.lower())
        if pid is None:
            return [], '', '', f'Profession not found: {name}'
        ids.append(pid)
        if not first_en:
            # Get the display name from lookups
            prof_obj = ProfessionMaster.query.get(pid)
            if prof_obj:
                first_en = prof_obj.name_en or ''
                first_ar = prof_obj.name_ar or ''

    return ids, first_en, first_ar, None


def build_employee(rowvals, lookups, seen_codes):
    """Validate one row and return (employee_data_dict, profession_ids_list)."""
    data = {}

    # Auto-generate code if blank
    code = clean_str(rowvals.get('employee_code'))
    if not code:
        code = generate_emp_code()

    name = clean_str(rowvals.get('name'))
    if not name:
        raise RowError('Employee Name is required')
    if rowvals.get('joining_date') in (None, ''):
        raise RowError('Joining Date is required')
    if code.lower() in seen_codes:
        raise RowError('Duplicate Employee Code within the Excel file')
    if code.lower() in lookups['existing_codes']:
        raise RowError('Employee Code already exists')

    data['employee_code'] = code
    data['name'] = name
    data['auto_code'] = parse_bool_cell(rowvals.get('auto_code'), 'Auto Code', default=False)
    data['is_active'] = parse_bool_cell(rowvals.get('is_active'), 'Active', default=True)
    data['is_muslim'] = parse_bool_cell(rowvals.get('is_muslim'), 'Muslim', default=False)

    for f in DATE_FLDS:
        hdr = next(h for h, fld in COLUMNS if fld == f)
        data[f] = parse_date_cell(rowvals.get(f), hdr)

    for f in NUM_FLDS:
        hdr = next(h for h, fld in COLUMNS if fld == f)
        data[f] = parse_number_cell(rowvals.get(f), hdr)

    data['email'] = validate_email(rowvals.get('email'))
    data['mobile'] = validate_mobile(rowvals.get('mobile'))

    # Handle professions (multi-select via comma-separated)
    prof_str = clean_str(rowvals.get('profession'))
    prof_ids, first_prof_en, first_prof_ar, prof_error = lookup_profession_ids(prof_str, lookups)
    if prof_error:
        raise RowError(prof_error)

    # Set single profession fields for backward compatibility (first profession)
    if prof_ids:
        data['profession_id'] = prof_ids[0]
        data['profession'] = first_prof_en
        data['profession_ar'] = first_prof_ar

    # Handle all other text fields
    handled = set(data.keys()) | {'buyer', 'employee_type'}
    for _, f in COLUMNS:
        if f in handled or f in DATE_FLDS or f in NUM_FLDS or f in BOOL_FLDS:
            continue
        if f in ('employee_type', 'profession', 'profession_ar', 'profession_id'):
            continue
        data[f] = clean_str(rowvals.get(f))

    data.pop('employee_type', None)

    # Overtime calculation: (basic / 30 * 8) * ratio
    basic = data.get('basic_salary', 0) or 0
    ratio = data.get('overtime_ratio', 0) or 0
    if basic > 0 and ratio > 0:
        data['overtime_rate'] = round(basic / 30 * 8 * ratio, 2)

    # Buyer lookup
    buyer_name = clean_str(rowvals.get('buyer'))
    if buyer_name:
        bid = lookups['buyers'].get(buyer_name.lower())
        if bid is None:
            raise RowError(f'Buyer not found: {buyer_name}')
        data['buyer_id'] = bid

    seen_codes.add(code.lower())
    return data, prof_ids


def load_lookups():
    """Load profession and buyer lookups + existing employee codes."""
    professions = {}
    for p in ProfessionMaster.query.all():
        if p.name_en:
            professions[p.name_en.strip().lower()] = p.id
        if p.name_ar:
            professions.setdefault(p.name_ar.strip().lower(), p.id)

    buyers = {}
    for b in BuyerMaster.query.all():
        if b.buyer_name_en:
            buyers[b.buyer_name_en.strip().lower()] = b.id
        if b.buyer_name_ar:
            buyers.setdefault(b.buyer_name_ar.strip().lower(), b.id)

    existing_codes = {c[0].lower() for c in db.session.query(Employee.employee_code).all() if c[0]}
    return {'professions': professions, 'buyers': buyers, 'existing_codes': existing_codes}


# ═══════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════

@emp_import_bp.route('/employees/import/template')
@login_required
@admin_required
def download_template():
    """Download sample Excel template with headers + example row."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employees'
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 3)
    ws.freeze_panes = 'A2'

    sample = {
        'Employee Code': '', 'Auto Code': 'No', 'Active': 'Yes', 'Muslim': 'Yes',
        'Employee Name': 'Ahmed Ali', 'Employee Name Arabic': 'أحمد علي',
        'Nationality': 'Pakistani', 'Profession': 'Welder',
        'Joining Date': '2026-01-15', 'Birth Date': '1990-05-20',
        'Mobile': '0501234567', 'Email': 'ahmed@example.com',
        'Salary Type': 'salary', 'Basic Salary': 3000, 'Working Hours': 8,
        'Overtime Ratio': 1.5, 'Work Status': 'active', 'Shift Type': 'day',
        'Passport Location': 'IN', 'PO Rate Unit': 'hour',
    }
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=2, column=i, value=sample.get(h, ''))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='employee_import_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@emp_import_bp.route('/employees/import', methods=['POST'])
@login_required
@admin_required
def import_employees():
    """Parse + validate + import Excel file. Returns JSON summary + failed rows."""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': _t('No file selected.', 'لم يتم اختيار ملف')}), 400

    ext = ('.' + f.filename.rsplit('.', 1)[-1].lower()) if '.' in f.filename else ''
    if ext not in ALLOWED_EXT:
        return jsonify({'ok': False,
                        'error': _t('Only .xlsx / .xls files are allowed.',
                                    'يسمح فقط بملفات .xlsx / .xls')}), 400

    try:
        wb = load_workbook(f, data_only=True, read_only=True)
    except Exception:
        return jsonify({'ok': False,
                        'error': _t('Could not read the Excel file.',
                                    'تعذر قراءة ملف الإكسل')}), 400

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return jsonify({'ok': False, 'error': _t('The file is empty.', 'الملف فارغ')}), 400

    col_field = {}
    for idx, h in enumerate(header_row):
        if h is None:
            continue
        fld = FIELD_BY_HDR.get(str(h).strip())
        if fld:
            col_field[idx] = fld

    lookups = load_lookups()
    seen_codes = set()
    imported = failed = skipped_empty = duplicates = 0
    failed_rows = []
    batch = []
    profession_map = {}  # employee_code -> list of profession_ids
    BATCH_SIZE = 200

    excel_row_no = 1
    for raw in rows_iter:
        excel_row_no += 1
        if raw is None or all(v is None or str(v).strip() == '' for v in raw):
            skipped_empty += 1
            continue

        rowvals = {fld: (raw[idx] if idx < len(raw) else None)
                   for idx, fld in col_field.items()}
        try:
            data, prof_ids = build_employee(rowvals, lookups, seen_codes)
            emp = Employee(created_by=current_user.id, **data)
            batch.append(emp)
            # Store profession IDs to link after commit
            profession_map[data['employee_code']] = prof_ids
            imported += 1
            if len(batch) >= BATCH_SIZE:
                db.session.add_all(batch)
                db.session.flush()
                # Save professions for this batch
                for b_emp in batch:
                    pids = profession_map.get(b_emp.employee_code, [])
                    for pid in pids:
                        db.session.add(EmployeeProfession(
                            employee_id=b_emp.id,
                            profession_id=pid
                        ))
                db.session.flush()
                batch = []
        except RowError as e:
            failed += 1
            reason = str(e)
            if 'already exists' in reason or 'Duplicate' in reason:
                duplicates += 1
            failed_rows.append({
                'row': excel_row_no,
                'employee_code': clean_str(rowvals.get('employee_code')),
                'name': clean_str(rowvals.get('name')),
                'error': reason,
                'raw': [raw[i] if i < len(raw) else None for i in range(len(header_row))],
            })
        except Exception as e:
            db.session.rollback()
            failed += 1
            failed_rows.append({
                'row': excel_row_no,
                'employee_code': clean_str(rowvals.get('employee_code')),
                'name': clean_str(rowvals.get('name')),
                'error': f'Database error: {e.__class__.__name__}',
                'raw': [raw[i] if i < len(raw) else None for i in range(len(header_row))],
            })

    try:
        if batch:
            db.session.add_all(batch)
        db.session.flush()
        # Save professions for remaining batch
        for b_emp in batch:
            pids = profession_map.get(b_emp.employee_code, [])
            for pid in pids:
                db.session.add(EmployeeProfession(
                    employee_id=b_emp.id,
                    profession_id=pid
                ))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        log.exception('Employee import commit failed')
        return jsonify({'ok': False,
                        'error': _t('Import failed at commit stage.',
                                    'فشل الاستيراد في مرحلة الحفظ')}), 500

    total = imported + failed + skipped_empty
    session['emp_import_failed'] = {
        'headers': [str(h) if h is not None else '' for h in header_row],
        'rows': [
            {'raw': [('' if v is None else (v.strftime('%Y-%m-%d') if isinstance(v, (datetime, date)) else str(v)))
                     for v in fr['raw']],
             'error': fr['error']}
            for fr in failed_rows
        ],
    }

    log.info('Employee import by user=%s file=%s total=%s imported=%s failed=%s',
             current_user.id, f.filename, total, imported, failed)

    return jsonify({
        'ok': True,
        'summary': {
            'total': total,
            'imported': imported,
            'failed': failed,
            'duplicates': duplicates,
            'skipped_empty': skipped_empty,
        },
        'failed_rows': [
            {'row': fr['row'], 'employee_code': fr['employee_code'],
             'name': fr['name'], 'error': fr['error']}
            for fr in failed_rows
        ],
    })


@emp_import_bp.route('/employees/import/failed')
@login_required
@admin_required
def download_failed():
    """Download failed rows as Excel with Error Reason column."""
    payload = session.get('emp_import_failed')
    if not payload or not payload.get('rows'):
        return jsonify({'ok': False,
                        'error': _t('No failed records available.',
                                    'لا توجد سجلات فاشلة')}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = 'Failed Records'
    headers = payload['headers'] + ['Error Reason']
    hdr_fill = PatternFill('solid', fgColor='B91C1C')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) + 3)
    ws.freeze_panes = 'A2'

    for r, fr in enumerate(payload['rows'], 2):
        for c, v in enumerate(fr['raw'], 1):
            ws.cell(row=r, column=c, value=v)
        ws.cell(row=r, column=len(headers), value=fr['error'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='employee_import_failed.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')